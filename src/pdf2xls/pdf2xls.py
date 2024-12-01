"""The main module with the program entry point.

Enter detailed module description here

Author: Name (mail)
"""

# ******************************************************************************
# Copyright (c) 2024, Achim Brunner
# License: BSD 3 Clause
# ******************************************************************************

# Imports **********************************************************************

import sys
import os
import logging
import argparse
import pdfplumber
from openpyxl import Workbook


try:
    from pdf2xls.version import __version__, __author__, __email__, __repository__, __license__
except ModuleNotFoundError:
    # provide dummy information when not installed as package but called directly
    # also necessary to get sphinx running without error
    __version__ = 'dev'
    __author__ = 'development'
    __email__ = 'none'
    __repository__ = 'none'
    __license__ = 'none'

# Variables ********************************************************************

LOG: logging.Logger = logging.getLogger(__name__)

# Classes **********************************************************************

# Functions ********************************************************************


def mm_to_points(mm: float) -> float:
    """
    Convert millimeters to points.
    1 mm = 2.83465 points

    Args:
        mm (float): Measurement in millimeters.

    Returns:
        float: Measurement in points.
    """
    return mm * 2.83465


def write_text_to_sheet(sheet, text_data: list):
    """
    Write text data to the given Excel sheet.
    Each item in the list is written to a new row.

    Args:
        sheet: The Excel worksheet where text will be written.
        text_data (List[str]): List of strings containing the text data to write.
    """
    for i, line in enumerate(text_data, start=1):
        sheet.cell(row=i, column=1, value=line)


def write_table_to_sheet(sheet, table: list):
    """
    Write a table (list of lists) to the given Excel sheet.
    Each sublist is written as a row.

    Args:
        sheet: The Excel worksheet where the table will be written.
        table (List[List[str]]): 2D list containing table data to write.
    """
    for row_idx, row in enumerate(table, start=1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)


def pdf_to_excel(input_file: str, output_file: str, start_page: int, end_page: int, header_height_mm: float, footer_height_mm: float):
    """
    Convert specified pages of a PDF to an Excel file, extracting text body and tables.
    The text body is saved in the first sheet, and each found table is saved in a new sheet.

    Args:
        input_file (str): Path to the input PDF file.
        output_file (str): Path to the output Excel file.
        start_page (int): Page number to start the conversion (1-based index).
        end_page (int): Page number to stop the conversion (inclusive, 1-based index).
        header_height_mm (float): Height of the header to ignore (in millimeters).
        footer_height_mm (float): Height of the footer to ignore (in millimeters).
    """
    try:
        # Validate input file
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"The input file '{input_file}' does not exist.")
        if not input_file.lower().endswith(".pdf"):
            raise ValueError("The input file must be a PDF.")

        # Validate output file
        if not output_file.lower().endswith(".xlsx"):
            raise ValueError("The output file must have an .xlsx extension.")

        # Convert header and footer heights from mm to points
        header_height = mm_to_points(header_height_mm)
        footer_height = mm_to_points(footer_height_mm)

        # Open the PDF file
        try:
            pdf = pdfplumber.open(input_file)
        except Exception as e:
            raise IOError(f"Failed to open the PDF file: {e}")

        # Create a new Excel workbook
        workbook = Workbook()

        # Add the first sheet for the text body
        text_sheet = workbook.active
        text_sheet.title = "Text_Body"
        all_text = []  # Collect text from all pages

        try:
            # Process each page in the range
            for i in range(start_page - 1, end_page):
                try:
                    page = pdf.pages[i]
                except IndexError:
                    raise IndexError(f"Page {i + 1} is out of range. The PDF has {len(pdf.pages)} pages.")

                # Get page height
                page_height = page.height

                # Define the cropping box to exclude headers and footers
                cropping_box = (0, header_height, page.width, page_height - footer_height)
                cropped_page = page.within_bbox(cropping_box)

                # Extract text and append to list
                page_text = cropped_page.extract_text()
                if page_text:
                    all_text.append(f"Page {i + 1}:\n{page_text}")

                # Extract tables and write to new sheets
                tables = cropped_page.extract_tables()
                for table_idx, table in enumerate(tables, start=1):
                    table_sheet = workbook.create_sheet(title=f"Table_Page_{i + 1}_{table_idx}")
                    write_table_to_sheet(table_sheet, table)

        except Exception as e:
            raise RuntimeError(f"Error processing PDF pages: {e}")

        finally:
            pdf.close()

        # Write all collected text to the text sheet
        write_text_to_sheet(text_sheet, all_text)

        # Save the workbook
        try:
            workbook.save(output_file)
        except Exception as e:
            raise IOError(f"Failed to save the Excel file: {e}")

        print(f"PDF data successfully converted to {output_file}")

    except FileNotFoundError as e:
        print(f"File error: {e}")
    except ValueError as e:
        print(f"Value error: {e}")
    except IndexError as e:
        print(f"Index error: {e}")
    except IOError as e:
        print(f"I/O error: {e}")
    except RuntimeError as e:
        print(f"Runtime error: {e}")


# Main *************************************************************************
def main() -> int:
    """ The program entry point function.

    Parsing and validating command line arguemts, to call the pdf processing according provided arugments

    Returns:
        int: System exit status.
    """
    logging.basicConfig(level=logging.INFO)

    # Create an ArgumentParser object
    parser = argparse.ArgumentParser(description="Convert PDF to Excel")

    # Add arguments
    parser.add_argument('-i', '--input_file', help='Path to the input PDF file', required=True)
    parser.add_argument('-o', '--output_file', help='Path to the output Excel file')
    parser.add_argument('-s', '--start_page', type=int, help='Page number to start the conversion')
    parser.add_argument('-e', '--end_page', type=int, help='Page number to stop the conversion')
    parser.add_argument('-hh', '--header_height_mm', type=float, help='Height of the header to ignore (in mm)')
    parser.add_argument('-fh', '--footer_height_mm', type=float, help='Height of the footer to ignore (in mm)')
    parser.add_argument('--verbose', action='store_true', help='Increase output verbosity')
    parser.add_argument('--version', action='version', version='%(prog)s' + __version__)

    args = parser.parse_args()  # Parse the arguments

    if args.verbose:
        print("Verbose mode is enabled.")
        LOG.setLevel(logging.DEBUG)

    pdf_to_excel(
        args.input_file,
        args.output_file,
        args.start_page,
        args.end_page,
        args.header_height_mm,
        args.footer_height_mm,
    )

    LOG.info("Hello World!")
    return 0  # return without errors


if __name__ == "__main__":
    sys.exit(main())
